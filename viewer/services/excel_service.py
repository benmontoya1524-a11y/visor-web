from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def build_xlsx_one_sheet(title: str, headers: list[str], rows: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(title)

    # Escribir headers
    ws.append(headers)

    # Escribir filas
    for r in rows:
        ws.append(list(r))

    # Aplicar formato 3 decimales a números
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"

    # Ajuste ancho columnas
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(40, len(str(h)) + 2))

    return wb


def safe_sheet_name(name: str) -> str:
    bad = ['\\', '/', '*', '?', ':', '[', ']']
    for b in bad:
        name = name.replace(b, "_")
    return name.strip()[:31] or "Sheet"
